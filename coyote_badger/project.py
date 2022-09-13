import os
import shutil

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from coyote_badger.config import CONVERTER_FOLDER_PREFIX, PROJECTS_FOLDER
from coyote_badger.source import Header, Kind, Source
from coyote_badger.utils import clean_string

SOURCE_SHEET = "Sources"
HEADER_ROW = 2
DATA_START_ROW = HEADER_ROW + 1


class Project(object):
    def __init__(self, name, xls_file=None):
        self.name = name
        self.project_folder = os.path.join(PROJECTS_FOLDER, name)
        self.project_folder_exists = os.path.isdir(self.project_folder)
        self.pull_folder = os.path.join(self.project_folder, "pull")
        self.pull_folder_exists = os.path.isdir(self.pull_folder)
        self.sources_file = os.path.join(self.project_folder, "Sources.xlsx")
        self.sources_file_exists = os.path.isfile(self.sources_file)

        # Create the data folders if the project doesn't exist
        if not self.pull_folder_exists:
            os.makedirs(self.pull_folder)

        # Create the sources file if the project doesn't exist
        if not self.sources_file_exists:
            xls_file.save(self.sources_file)

        self.wb = load_workbook(self.sources_file)
        self.ws = self.wb[SOURCE_SHEET]
        self.headers = self.ws[HEADER_ROW]
        self._header_index = None

        # If we are creating this project for the first time, clean it
        if xls_file:
            self.clean_wb()

    @property
    def header_index(self):
        """A property containing a mapping of our known
        Headers to what column index they are in the Excel
        sources sheet.

        Each column index is 1-indexed (e.g. Column A is 1, not 0).
        :returns: Mapping of Header value to column index
        :rtype: {dict(str -> int)}
        """
        if self._header_index is not None:
            return self._header_index
        index = {}
        for cell in self.headers:
            if cell.value in set(item.value for item in Header):
                index[cell.value] = cell.column  # column is 1-indexed
        self._header_index = index
        return self._header_index

    def delete(self):
        """Deletes a project.

        Removes the project foler, along with it's pulled source
        pdfs and it's Sources.xlsx file.
        """
        shutil.rmtree(self.project_folder, ignore_errors=True)

    @staticmethod
    def get_projects():
        """Gets all of the projects in the data directory.

        :returns: A list of project names
        :rtype: {[str]}
        """
        projects = []
        for item in os.listdir(PROJECTS_FOLDER):
            if os.path.isdir(
                os.path.join(PROJECTS_FOLDER, item)
            ) and not item.startswith(CONVERTER_FOLDER_PREFIX):
                projects.append(item)
        return projects

    @staticmethod
    def get_project(name):
        """Gets a project by name.

        An alternative way of getting a project by its name.
        :param name: The name of the project to get
        :type name: str
        :returns: The project
        :rtype: {Project}
        """
        if name in Project.get_projects():
            return Project(name)
        return None

    def get_sources(self):
        """Gets all the sources in the Sources.xlsx.

        :returns: A list of the sources
        :rtype: {[Source]}
        """
        sources = []
        for row in self.ws.iter_rows(min_row=DATA_START_ROW):
            source_index = row[0].row - HEADER_ROW
            source = self.get_source(source_index)
            sources.append(source)
        return sources

    def get_source(self, index):
        """Gets a single source from the Sources.xlsx file.

        :param index: The row the source is at (1-indexed)
        :type index: int
        :returns: The source
        :rtype: {Source}
        """
        row = self.ws[HEADER_ROW + index]
        source = self.build_source_from_row(row)
        return source

    def save_sources(self, sources):
        """Saves all the provided sources back to the Sources.xlsx file.

        :param sources: The sources to save
        :type sources: [Sources]
        """
        for i, source in enumerate(sources):
            self.save_source(i + 1, source)
        self.wb.save(self.sources_file)

    def save_source(self, index, source):
        """Saves a single source back to the Sources.xslx file.

        :param index: The row of the source to save (1-indexed)
        :type index: int
        :param source: The source to save
        :type source: Source
        """
        row = self.ws[HEADER_ROW + index]
        self.build_row_from_source(row, source)
        self.wb.save(self.sources_file)

    def save_pull_path(self, filename, extension=None):
        """The path to save a pulled resource at for this project.

        :param filename: The name of the file to be saved
        :type filename: str
        :param extension: The extension of the file, defaults to None
        :type extension: str, optional
        :returns: The path the file should be saved at
        :rtype: {str}
        """
        if extension:
            filename = f"{filename}.{extension}"
        return os.path.join(self.pull_folder, filename)

    def build_source_from_row(self, row):
        """Builds a source given a row.

        :param row: The row in the excel
        :type row: [Cell]
        :returns: The source that was generated from this data
        :rtype: {Source}
        """
        values = [c.value for c in row]
        return Source(
            fn_num=values[self.header_index[Header.fn_num.value] - 1],
            long_cite=values[self.header_index[Header.long_cite.value] - 1],
            short_cite=values[self.header_index[Header.short_cite.value] - 1],
            filename=values[self.header_index[Header.filename.value] - 1],
            library=values[self.header_index[Header.library.value] - 1],
            has_book=values[self.header_index[Header.has_book.value] - 1],
            kind=values[self.header_index[Header.kind.value] - 1],
            result=values[self.header_index[Header.result.value] - 1],
        )

    def build_row_from_source(self, row, source):
        """Fills a worksheet's row with data from a provided source.

        :param row: The previous row to fill in
        :type row: [Cell]
        :param source: The source to use for data
        :type source: Source
        :returns: The new row values
        :rtype: {[Cell]}
        """
        for cell in row:
            if cell.col_idx == self.header_index[Header.fn_num.value]:
                cell.value = source.fn_num or cell.value
                cell.number_format = "0.00"
            elif cell.col_idx == self.header_index[Header.long_cite.value]:
                cell.value = source.long_cite or cell.value
                cell.alignment = Alignment(wrap_text=True)
            elif cell.col_idx == self.header_index[Header.short_cite.value]:
                if source.kind == Kind.WEBSITE:
                    cell.hyperlink = source.short_cite
                    cell.style = "Hyperlink"
                cell.value = source.short_cite or cell.value
                cell.alignment = Alignment(wrap_text=True)
            elif cell.col_idx == self.header_index[Header.filename.value]:
                cell.value = source.filename or cell.value
                cell.alignment = Alignment(wrap_text=True)
            elif cell.col_idx == self.header_index[Header.kind.value]:
                cell.value = source.kind.value or cell.value
            elif cell.col_idx == self.header_index[Header.result.value]:
                cell.value = source.result.value or cell.value
        return row

    def clean_wb(self):
        """Cleans a workbook.

        Removes extraneous characters, empty rows,
        and None values from a Sources workbook.
        """
        delete_rows = []
        for row in self.ws.iter_rows(min_row=DATA_START_ROW):
            # Clean the strings
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = cell.value or ""
                    cell.value = clean_string(cell.value)
            # Keep track of blank rows
            values = [cell.value for cell in row]
            if not any(values):
                delete_rows.append(row[0].row)
        # Delete blank rows
        for index in reversed(delete_rows):
            self.ws.delete_rows(index)
        self.wb.save(self.sources_file)
